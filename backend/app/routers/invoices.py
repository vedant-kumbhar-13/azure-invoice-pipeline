"""
Invoice router for InvoiceAI.

VUL-01: Stores blob_name, returns file_url_sas on demand.
VUL-03: Background task no longer holds file_bytes.
BUG-07: Uses unified processing pipeline.
BUG-09: Search queries JSON fields (vendor_name, invoice_number, GSTIN).
BUG-14: CSV export uses streaming generator; XLSX capped at 10,000 rows.
"""
import hashlib
import logging
import io
import csv
import asyncio
import uuid
from typing import List

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, status, Header
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func as sqlfunc, or_, cast, String
from app.database import get_db, SessionLocal
from app.models.user import User
from app.models.invoice import Invoice
from app.middleware.auth import get_current_user
from app.middleware.rate_limit import rate_limited_user
from app.services.blob_storage import upload_file_to_blob, get_blob_sas_url, delete_blob
from app.utils.datetime_utils import utc_iso
from app.services.azure_ai import analyze_invoice
from app.services.invoice_mapper import map_fields, map_gst_qr_to_canonical

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s [%(name)s] %(message)s")

router = APIRouter(prefix="/invoices", tags=["Invoices"])

ALLOWED_EXTENSIONS = {'.pdf', '.jpg', '.jpeg', '.png'}
MAX_FILE_SIZE = 20 * 1024 * 1024  # 20 MB

# ── File magic bytes signatures ─────────────────
FILE_SIGNATURES = {
    '.pdf':  [b'%PDF'],
    '.jpg':  [b'\xff\xd8\xff'],
    '.jpeg': [b'\xff\xd8\xff'],
    '.png':  [b'\x89PNG'],
}


def _validate_file_magic(file_bytes: bytes, ext: str) -> bool:
    """Verify actual file content matches the claimed extension."""
    sigs = FILE_SIGNATURES.get(ext, [])
    if not sigs:
        return True  # No known signature, allow
    return any(file_bytes[:len(sig)] == sig for sig in sigs)


# ── BUG-07 + BUG-B5: Bounded executor ──────────────────────────────────────
# The thread pool lives in app.main._pipeline_executor.  We lazy-import it
# inside the functions that submit work to avoid a circular import
# (app.main → app.routers.invoices → app.main).

logger = logging.getLogger("invoiceai.invoices")


def _get_executor():
    """Lazy accessor that avoids the circular import at module level."""
    from app.main import _pipeline_executor
    return _pipeline_executor


def _run_full_task(invoice_id: str, file_bytes: bytes, filename: str):
    """Background task: uploads blob, saves URL, then runs the full pipeline.

    Runs inside _pipeline_executor so the HTTP response is not blocked.
    Uses its own DB session (the request session is long gone by now).
    """
    from app.services.processing_pipeline import run_invoice_pipeline
    db = SessionLocal()
    try:
        # a) Upload to blob
        blob_name = upload_file_to_blob(file_bytes, filename)

        # b) Persist blob_name in the invoice record
        invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
        if not invoice:
            logger.error(f"[full_task] Invoice {invoice_id} not found in DB — aborting")
            return
        invoice.file_url = blob_name
        db.commit()

        # c) Generate a time-limited SAS URL for Azure Document Intelligence
        blob_url = get_blob_sas_url(blob_name)

        # d) Run the unified processing pipeline
        run_invoice_pipeline(invoice_id, blob_url, filename)
    except Exception as exc:
        # e) On ANY failure — mark invoice as failed with the error detail
        logger.exception(f"[full_task] Invoice {invoice_id} failed: {exc}")
        try:
            invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
            if invoice:
                invoice.status = "failed"
                invoice.error_detail = str(exc)[:500]
                db.commit()
        except Exception as db_err:
            logger.error(f"[full_task] Could not update failure status for {invoice_id}: {db_err}")
    finally:
        db.close()
        # Free potentially large buffer
        del file_bytes


def _run_pipeline_task(invoice_id: str, blob_name: str, filename: str):
    """Background task wrapper for *reprocess*: blob already uploaded.

    Generates a fresh SAS URL and runs the unified pipeline.
    """
    from app.services.processing_pipeline import run_invoice_pipeline
    blob_url = get_blob_sas_url(blob_name)
    run_invoice_pipeline(invoice_id, blob_url, filename)


@router.post("/upload", status_code=status.HTTP_202_ACCEPTED)
async def upload_invoice(
    file: UploadFile = File(...),
    x_idempotency_key: str = Header(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(rate_limited_user),
):
    import os

    # Handle idempotency
    hashed_key = None
    if x_idempotency_key:
        hashed_key = hashlib.sha256(x_idempotency_key.encode("utf-8")).hexdigest()
        existing_invoice = db.query(Invoice).filter(
            Invoice.user_id == current_user.id,
            Invoice.idempotency_key == hashed_key,
        ).first()

        if existing_invoice:
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "id": existing_invoice.id,
                    "status": existing_invoice.status,
                    "source_type": existing_invoice.source_type or "UNKNOWN",
                    "ingestion_method": existing_invoice.ingestion_method or "PENDING",
                },
            )

    # Extension check
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"File type '{ext}' not allowed. Use PDF, JPG, or PNG.",
        )

    file_bytes = await file.read()
    if len(file_bytes) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="File exceeds 20MB limit.")

    # Magic bytes validation (Fix #9)
    if not _validate_file_magic(file_bytes, ext):
        raise HTTPException(
            status_code=400,
            detail="File content does not match the claimed extension. Possible corrupted or disguised file.",
        )

    from sqlalchemy.exc import IntegrityError

    new_invoice = Invoice(
        user_id=current_user.id,
        status="processing",
        original_filename=file.filename,
        idempotency_key=hashed_key,
        source_type="UNKNOWN",
        ingestion_method="PENDING",
    )

    # BUG-08: Handle race condition on idempotency key
    try:
        db.add(new_invoice)
        db.commit()
        db.refresh(new_invoice)
    except IntegrityError:
        db.rollback()
        existing = db.query(Invoice).filter(
            Invoice.user_id == current_user.id,
            Invoice.idempotency_key == hashed_key,
        ).first()
        if existing:
            return JSONResponse(
                status_code=200,
                content={"id": existing.id, "status": existing.status},
            )
        raise HTTPException(status_code=500, detail="Unexpected conflict during upload.")

    # Submit blob upload + pipeline to the bounded thread pool.
    # Both the blocking upload_file_to_blob() (5–15 s) and the pipeline
    # now run in the background so the HTTP response returns immediately.
    _get_executor().submit(
        _run_full_task,
        new_invoice.id,
        file_bytes,
        file.filename,
    )

    # Free the request-scoped reference; the background thread has its own copy.
    del file_bytes

    return {
        "id": new_invoice.id,
        "status": new_invoice.status,
        "source_type": new_invoice.source_type,
        "ingestion_method": new_invoice.ingestion_method,
    }


@router.get("/stats")
def get_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Fix #8: Count all final-positive statuses as 'completed',
    and all review statuses as one number."""
    total = db.query(Invoice).filter(Invoice.user_id == current_user.id).count()

    completed = db.query(Invoice).filter(
        Invoice.user_id == current_user.id,
        Invoice.status.in_(["completed", "AUTO_APPROVED", "VERIFIED"]),
    ).count()

    processing = db.query(Invoice).filter(
        Invoice.user_id == current_user.id,
        Invoice.status.in_(["processing", "NEEDS_REVIEW", "HUMAN_REQUIRED"]),
    ).count()

    failed = db.query(Invoice).filter(
        Invoice.user_id == current_user.id,
        Invoice.status.in_(["failed", "REJECTED"]),
    ).count()

    avg_conf = db.query(sqlfunc.avg(Invoice.confidence)).filter(
        Invoice.user_id == current_user.id,
        Invoice.status.in_(["completed", "AUTO_APPROVED", "VERIFIED"]),
    ).scalar()

    return {
        "total": total,
        "completed": completed,
        "processing": processing,
        "failed": failed,
        "avg_confidence": round(float(avg_conf or 0), 4),
    }


# BUG-18: Safe extraction from data_json that handles non-dict or corrupted values
def _safe_get(data_json, field):
    if not isinstance(data_json, dict):
        return None
    field_data = data_json.get(field, {})
    if isinstance(field_data, dict):
        return field_data.get("value")
    return None


@router.get("/")
def list_invoices(
    skip: int = 0,
    limit: int = 20,
    status: str = None,
    search: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # BUG-17: Server-side filtering instead of client-side over 1000 records
    query = db.query(Invoice).filter(Invoice.user_id == current_user.id)

    if status and status != "All":
        status_map = {
            "Processing": ["processing"],
            "Auto-Approved": ["AUTO_APPROVED", "completed", "VERIFIED"],
            "Needs Review": ["NEEDS_REVIEW", "HUMAN_REQUIRED"],
            "Failed": ["failed", "REJECTED"],
        }
        status_values = status_map.get(status, [status])
        query = query.filter(Invoice.status.in_(status_values))

    # BUG-09: Search across vendor name, invoice number, GSTIN, and filename
    if search and search.strip():
        search_term = f"%{search.strip()}%"
        from app.config import settings
        if settings.DATABASE_URL.startswith("postgresql"):
            # PostgreSQL: Use JSON path operators for precise field search
            vendor_search = Invoice.data_json["vendor_name"]["value"].astext.ilike(search_term)
            inv_num_search = Invoice.data_json["invoice_number"]["value"].astext.ilike(search_term)
            gstin_search = Invoice.data_json["vendor_gstin"]["value"].astext.ilike(search_term)
            filename_search = Invoice.original_filename.ilike(search_term)
            query = query.filter(or_(vendor_search, inv_num_search, gstin_search, filename_search))
        else:
            # SQLite: Cast JSON to text and use broad LIKE search
            json_text = cast(Invoice.data_json, String)
            query = query.filter(or_(
                Invoice.original_filename.ilike(search_term),
                json_text.ilike(search_term),
            ))

    total = query.count()

    invoices = (
        query
        .order_by(Invoice.created_at.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )

    return {
        "items": [
            {
                "id": inv.id,
                "status": inv.status,
                "original_filename": inv.original_filename,
                "created_at": utc_iso(inv.created_at),
                "confidence_score": inv.confidence,
                "vendor_name": _safe_get(inv.data_json, "vendor_name"),
                "total_amount": _safe_get(inv.data_json, "total_amount"),
                "invoice_number": _safe_get(inv.data_json, "invoice_number"),
                "ingestion_method": inv.ingestion_method,
                "source_type": inv.source_type,
                # VUL-01: file_url removed from list view — no direct file access needed
            }
            for inv in invoices
        ],
        "total": total,
        "search_fields": ["filename", "vendor_name", "invoice_number", "vendor_gstin"],
    }


@router.get("/{invoice_id}")
def get_invoice(
    invoice_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.user_id == current_user.id,
    ).first()

    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found.")

    # VUL-01: Generate fresh SAS URL on demand from stored blob_name
    file_url_sas = None
    if invoice.file_url:
        try:
            file_url_sas = get_blob_sas_url(invoice.file_url)
        except Exception:
            pass

    return {
        "id": invoice.id,
        "status": invoice.status,
        "original_filename": invoice.original_filename,
        # BUG-C3: blob_name (file_url) is an internal implementation detail — never expose it.
        # Consumers only need the time-limited SAS URL.
        "file_url_sas": file_url_sas,
        "created_at": utc_iso(invoice.created_at),
        "confidence_score": invoice.confidence,
        "data": invoice.data_json,
        "data_json": invoice.data_json,
        "source_type": invoice.source_type,
        "ingestion_method": invoice.ingestion_method,
        "gst_rules_json": invoice.gst_rules_json,
        "processing_time_ms": invoice.processing_time_ms,
        "error_message": invoice.error_detail if invoice.status in ("failed", "HUMAN_REQUIRED") else None,
    }


@router.delete("/{invoice_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_invoice(
    invoice_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.user_id == current_user.id,
    ).first()

    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found.")

    # BUG-A3: Delete the Azure blob before removing the DB record.
    # Try/except ensures a transient storage error is logged but does NOT block erasure.
    if invoice.file_url:
        try:
            delete_blob(invoice.file_url)
        except Exception as blob_err:
            logger.warning(
                f"[invoice] Blob deletion failed for {invoice.file_url} "
                f"(invoice {invoice_id}): {blob_err} — continuing with DB delete."
            )

    db.delete(invoice)
    db.commit()


@router.post("/{invoice_id}/reprocess", status_code=status.HTTP_202_ACCEPTED)
def reprocess_invoice(
    invoice_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """BUG-07: Uses unified pipeline (includes QR detection on reprocess)."""
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.user_id == current_user.id,
    ).first()

    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found.")

    if invoice.status == "processing":
        raise HTTPException(status_code=400, detail="Invoice is already being processed.")

    if not invoice.file_url:
        raise HTTPException(status_code=400, detail="No file URL stored. Cannot reprocess.")

    invoice.status = "processing"
    invoice.error_detail = None
    db.commit()

    # BUG-07 + BUG-B5: Submit to the bounded pipeline thread pool.
    _get_executor().submit(
        _run_pipeline_task,
        invoice.id,
        invoice.file_url,       # VUL-01: file_url is blob_name
        invoice.original_filename or "unknown.pdf",
    )

    return {"id": invoice.id, "status": "processing"}


# ── Bulk Upload Endpoints ──────────────────────────────────────────────────

MAX_BATCH_FILES = 20

# Terminal statuses for determining if batch processing is complete
_TERMINAL_STATUSES = {
    "AUTO_APPROVED", "NEEDS_REVIEW", "HUMAN_REQUIRED",
    "failed", "VERIFIED", "REJECTED", "completed",
}


@router.post("/upload/bulk", status_code=status.HTTP_202_ACCEPTED)
async def upload_invoices_bulk(
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(rate_limited_user),
):
    """Bulk upload up to 20 invoice files in one request.

    Each file is validated independently.  Valid files get an Invoice DB
    record and are submitted to the background thread pool.  Invalid
    files are collected in the 'rejected' list.
    """
    import os

    if len(files) > MAX_BATCH_FILES:
        raise HTTPException(
            status_code=400,
            detail=f"Max {MAX_BATCH_FILES} files per batch.",
        )

    batch_id = str(uuid.uuid4())
    accepted_invoices: list[dict] = []
    rejected_files: list[dict] = []

    for file in files:
        filename = file.filename or "unknown"
        ext = os.path.splitext(filename)[1].lower()

        # Extension check
        if ext not in ALLOWED_EXTENSIONS:
            rejected_files.append({
                "filename": filename,
                "reason": f"File type '{ext}' not allowed. Use PDF, JPG, or PNG.",
            })
            continue

        file_bytes = await file.read()

        # Size check
        if len(file_bytes) > MAX_FILE_SIZE:
            rejected_files.append({"filename": filename, "reason": "File exceeds 20MB limit."})
            del file_bytes
            continue

        # Magic bytes check
        if not _validate_file_magic(file_bytes, ext):
            rejected_files.append({
                "filename": filename,
                "reason": "File content does not match the claimed extension.",
            })
            del file_bytes
            continue

        # Create Invoice DB record
        new_invoice = Invoice(
            user_id=current_user.id,
            status="processing",
            original_filename=filename,
            batch_id=batch_id,
            source_type="UNKNOWN",
            ingestion_method="PENDING",
        )
        db.add(new_invoice)
        db.commit()
        db.refresh(new_invoice)

        # Submit to background thread pool
        _get_executor().submit(
            _run_full_task,
            new_invoice.id,
            file_bytes,
            filename,
        )
        del file_bytes

        accepted_invoices.append({
            "invoice_id": new_invoice.id,
            "filename": filename,
            "status": new_invoice.status,
        })

    return {
        "batch_id": batch_id,
        "total_files": len(files),
        "accepted": len(accepted_invoices),
        "rejected": rejected_files,
        "invoices": accepted_invoices,
    }


@router.get("/batch/{batch_id}")
def get_batch_status(
    batch_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return live processing status for every invoice in a batch."""
    invoices = (
        db.query(Invoice)
        .filter(
            Invoice.batch_id == batch_id,
            Invoice.user_id == current_user.id,
        )
        .order_by(Invoice.created_at.asc())
        .all()
    )

    if not invoices:
        raise HTTPException(status_code=404, detail="Batch not found.")

    statuses = [inv.status for inv in invoices]
    completed_count = sum(1 for s in statuses if s in _TERMINAL_STATUSES)
    failed_count = sum(1 for s in statuses if s in ("failed", "REJECTED"))

    if all(s in _TERMINAL_STATUSES for s in statuses):
        overall = "COMPLETED"
    elif any(s == "processing" for s in statuses):
        overall = "PROCESSING"
    else:
        overall = "PARTIAL"

    return {
        "batch_id": batch_id,
        "overall_status": overall,
        "total": len(invoices),
        "completed": completed_count,
        "failed": failed_count,
        "invoices": [
            {
                "id": inv.id,
                "original_filename": inv.original_filename,
                "status": inv.status,
                "confidence_score": inv.confidence,
                "ingestion_method": inv.ingestion_method,
            }
            for inv in invoices
        ],
    }


# ── Export Endpoints ────────────────────────────
# BUG-14: Streaming CSV generator + XLSX 10K cap

def _csv_row(inv: Invoice) -> list:
    """Build a CSV row from an Invoice object."""
    data = inv.data_json or {}
    confidence_val = inv.confidence if inv.confidence is not None else 0.0

    def _g(key):
        v = data.get(key, {})
        return v.get("value", "") if isinstance(v, dict) else ""

    return [
        inv.id,
        inv.original_filename or "",
        inv.status or "",
        round(confidence_val, 4),
        _g("vendor_name"),
        _g("vendor_gstin"),
        _g("invoice_number"),
        _g("invoice_date"),
        _g("total_amount"),
        inv.created_at.isoformat() if inv.created_at else "",
        inv.processing_time_ms if inv.processing_time_ms is not None else "",
    ]


CSV_HEADERS = [
    "id", "original_filename", "status", "confidence_score",
    "vendor_name", "vendor_gstin", "invoice_number", "invoice_date",
    "total_amount", "created_at", "processing_time_ms",
]


@router.get("/export/csv")
def export_invoices_csv(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """BUG-14: Streaming CSV export — processes in batches of 500."""
    def _generate():
        # Header
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(CSV_HEADERS)
        yield buf.getvalue()

        offset = 0
        batch_size = 500
        while True:
            batch = (
                db.query(Invoice)
                .filter(Invoice.user_id == current_user.id)
                .order_by(Invoice.created_at.desc())
                .offset(offset)
                .limit(batch_size)
                .all()
            )
            if not batch:
                break

            buf = io.StringIO()
            writer = csv.writer(buf)
            for inv in batch:
                writer.writerow(_csv_row(inv))
            yield buf.getvalue()

            offset += batch_size

    return StreamingResponse(
        _generate(),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="invoices.csv"'},
    )


@router.get("/export/xlsx")
def export_invoices_xlsx(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """BUG-14: XLSX export with 10,000 row safety cap."""
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.comments import Comment

    MAX_XLSX_ROWS = 10_000

    invoices = (
        db.query(Invoice)
        .filter(Invoice.user_id == current_user.id)
        .order_by(Invoice.created_at.desc())
        .limit(MAX_XLSX_ROWS)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = [
        "ID", "Original Filename", "Status", "Confidence Score",
        "Vendor Name", "Vendor GSTIN", "Invoice Number", "Invoice Date",
        "Total Amount", "Created At", "Processing Time (ms)",
    ]

    ws.append(headers)

    # BUG-14: Add note if export was capped
    if len(invoices) >= MAX_XLSX_ROWS:
        ws["A1"].comment = Comment(
            f"Export limited to {MAX_XLSX_ROWS:,} records. Use CSV for full export.",
            "InvoiceAI"
        )

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    ws.freeze_panes = "A2"

    for inv in invoices:
        ws.append(_csv_row(inv))

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = min(adjusted_width, 50)

    b = io.BytesIO()
    wb.save(b)
    b.seek(0)

    return StreamingResponse(
        b,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="invoices.xlsx"'},
    )